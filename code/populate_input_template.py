"""Populate the canonical on/off input template from a legacy annual workbook.

Usage:
    python code/populate_input_template.py \
        data/awc_vwr_dtw_MASTER_05_2026.xlsx \
        data/onoff_input_05_2026.xlsx

The annual workbook still carries some legacy column names (for example,
Date_Julian). The canonical workbook intentionally drops those columns and uses
real dates only.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT / "data" / "input_template_2026.xlsx"


CANONICAL_SHEETS = {
    "awc_vwr": ["site", "date", "awc", "vwr"],
    "dtw": ["site", "date", "dtw"],
    "on_off_history": ["site", "date", "status", "status_code"],
    "current_status": ["site", "current_status", "awc_req_turnon"],
}


SOURCE_MAPS = {
    "awc_vwr": {
        "source_sheet": "awc_vwr",
        "columns": {
            "site": ["site", "Site"],
            "date": ["date", "Date", "date_real"],
            "awc": ["awc", "Adj AWC_soil"],
            "vwr": ["vwr", "VWR_veg"],
        },
    },
    "dtw": {
        "source_sheet": "dtw",
        "columns": {
            "site": ["site", "Site"],
            "date": ["date", "Date", "date_real"],
            "dtw": ["dtw", "DTW_BGS"],
        },
    },
    "on_off_history": {
        "source_sheet": "on.off",
        "columns": {
            "site": ["site", "Site"],
            "date": ["date", "Date", "date_real"],
            "status": ["on.off"],
            "status_code": ["on.off.1"],
        },
    },
    "current_status": {
        "source_sheet": "on.off.table",
        "columns": {
            "site": ["site"],
            "current_status": ["current.status"],
            "awc_req_turnon": ["AWC.req.turnon"],
        },
    },
}


def normalize_header(value: object) -> str:
    return "" if value is None else str(value).strip()


def header_index(ws) -> dict[str, int]:
    headers = [normalize_header(cell.value) for cell in ws[1]]
    return {header: idx for idx, header in enumerate(headers) if header}


def find_column(headers: dict[str, int], candidates: Iterable[str]) -> int:
    for candidate in candidates:
        if candidate in headers:
            return headers[candidate]
    raise KeyError(f"Missing column. Tried: {', '.join(candidates)}")


def read_rows(source_wb, sheet_name: str, column_map: dict[str, list[str]]) -> list[list[object]]:
    ws = source_wb[sheet_name]
    headers = header_index(ws)
    indices = [find_column(headers, column_map[column]) for column in column_map]
    rows: list[list[object]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        selected = [row[idx] for idx in indices]
        if any(value is not None for value in selected):
            rows.append(selected)

    return rows


def copy_static_sheet(template_wb, output_wb, sheet_name: str) -> None:
    source_ws = template_wb[sheet_name]
    target_ws = output_wb.create_sheet(sheet_name)

    for row in source_ws.iter_rows(values_only=True):
        target_ws.append(list(row))


def add_sheet(output_wb, sheet_name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = output_wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    for column in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=column, max_col=column, values_only=True):
            value = row[0]
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(column)].width = min(max(max_len + 2, 12), 45)

    if ws.max_row >= 1 and ws.max_column >= 1 and ws.title != "README":
        ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
        table = Table(displayName="tbl_" + ws.title.replace(".", "_"), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def build(source_path: Path, output_path: Path) -> None:
    source_wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    template_wb = openpyxl.load_workbook(TEMPLATE_PATH, read_only=True, data_only=True)
    output_wb = Workbook()
    output_wb.remove(output_wb.active)

    for sheet_name in ["README", "site_metadata", "linked_wells"]:
        copy_static_sheet(template_wb, output_wb, sheet_name)

    for sheet_name, headers in CANONICAL_SHEETS.items():
        source_spec = SOURCE_MAPS[sheet_name]
        rows = read_rows(
            source_wb,
            source_spec["source_sheet"],
            source_spec["columns"],
        )
        add_sheet(output_wb, sheet_name, headers, rows)

    for ws in output_wb.worksheets:
        style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_wb.save(output_path)
    source_wb.close()
    template_wb.close()


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python code/populate_input_template.py <source.xlsx> <output.xlsx>")

    build(Path(sys.argv[1]), Path(sys.argv[2]))


if __name__ == "__main__":
    main()
